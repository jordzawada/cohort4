from openpyxl import Workbook
from openpyxl import load_workbook


# def import():
# this function will have a bunch of searches that looks in a variety of files. the goal is to match tabs and columns and turn them into normalized rows. Run the verify function after this to make sure the scraped data is good. 



def createWorkbook():
    wb = Workbook()
    ws = wb.active
    ws.title= "customers"
    ws['A1'] = "id"
    ws['B1'] = "name"
    ws['A2'] = 1
    ws['B2'] = 'Al'

    ws = wb.create_sheet("invoices")
    ws['A1'] = "id"
    ws['B1'] = "customer id"
    ws['C1'] = "date"
    ws['A2'] =  1
    ws['A3'] = 2
    ws['B2'] = 1
    ws['B3'] = 1

    ws = wb.create_sheet("line items")
    ws['A1'] = "id"
    # ws['B1'] = "name" is this needed?
    ws['B1'] = "qty"
    ws['C1'] = "invoice id"
    ws['D1'] = "product id"
    ws['A2'] = 1
    ws['B2'] =50
    ws['C2'] = 1
    ws['D2'] =2

    ws['A3'] = 2
    ws['B3'] = 25
    ws['C3'] = 1
    ws['D3'] = 1

    ws = wb.create_sheet("product")
    ws['A1'] = "id"
    ws['B1'] = "name"
    ws['A2'] = 1
    ws['B2'] = 'bolts'
    ws['A3'] = 2
    ws['B3'] = 'nuts'
    wb.save('billy.xlsx')

def createInvoice(invNum):
    # print('invoice')
    wb=load_workbook('billy.xlsx')
    for row in wb['invoices']['A']:
        if row.value == invNum:
            customerID = wb['invoices'].cell(row=row.row, column=row.column+1).value
    for row in wb['customers']['A']:
        if row.value == customerID:
            customerName = wb['customers'].cell(row=row.row, column=row.column+1).value
    lines= {}        
    for row in wb['line items']['A']: 
        x = row.value  
        for invoiceID in wb['product']['C']:
            if x == productID.value:
                x = wb['product'].cell(row=productID.row,column=productID.column+1).value
        lines[x] = wb['line items'].cell(row=row.row, column=row.column+1).value
        # print(lines)  
    f = open('output.txt', 'w')
    f.write(f'Invoice {invNum} for customer:{customerName}, has items:' )
    for x in lines:
        f.write(x)
        f.write(str(lines[x]))           
    
# createWorkbook()

def validateBilly():
    wb=load_workbook('billy.xlsx')
    elements = [];
    for a in wb['product']['A']:
        elements.append(a.value)
        # add in int checker?
        if len(elements) > len(set(elements)): 
            print('Product ID cannot be duplicate,check ended early')
            return
    for name in wb['product']['B']:
        if type(name.value) != str:
            print ('Products name must be string')
    for price in wb['product']['C']:
        if price.value == 'price':
            continue
        elif price.value < 0 and type(price.value) != int:
            print('Prodcucts: invalid price entered')
            

    customerIDlist=[];
    for customerID in wb['customers']['A']:
        customerIDlist.append(customerID.value)
        if len(customerIDlist) > len(set(customerIDlist)): 
            print('Customer ID cannot be duplicate. Check ended early')
            return

    elements = [];
    a=0
    for invoiceID in wb['invoices']['A']:
        a=a+1
        elements.append(invoiceID.value)
        # add in int checker?
        if len(elements) > len(set(elements)): 
            print(f'Invoice ID cannot be duplicate, check ended early near line {a}, check for duplicate {elements[-2]}')
            return            
    for invoiceNum in wb['invoices']['B']:
        if invoiceNum.value == 'customer id':
            continue
        if type(invoiceNum.value) != int:
            print('Customer ID must be type int')
        if invoiceNum.value not in customerIDlist:
            print('Customer ID in invoices does not correspond to customer list')
        
validateBilly();