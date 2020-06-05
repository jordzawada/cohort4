import workbook
import os
from openpyxl import load_workbook
# os.system('python workbook.py')

wb=load_workbook('billy.xlsx')
# print(os.listdir())

def test_answer():
    if "billy.xlsx" in os.listdir():
        x = True
    assert x == True
    # print(wb.sheetnames)
    assert wb.sheetnames[0] == 'customers'
    assert wb['customers']['A1'].value == 'id'
    assert wb['customers']['B1'].value == 'name'
    assert wb['customers']['A2'].value == 1
    assert wb['customers']['B2'].value == 'Al'

    assert wb.sheetnames[1] == 'invoices'
    assert wb['invoices']['A1'].value == 'id'
    assert wb['invoices']['A2'].value == 1
    assert wb['invoices']['A3'].value == 2

    assert wb.sheetnames[2] =='line items'
    assert wb['line items']['B1'].value == 'qty'

    assert wb.sheetnames[3] == 'product'
    assert wb['product']['B1'].value == 'name'

    loaded= workbook.createInvoice(1)  